from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from typing import List, Dict, Any
from pathlib import Path
from app.config import settings
import json


def export_prompts_to_excel(
    pages_data: List[Dict[str, Any]],
    project_name: str
) -> str:
    """
    导出提示词到Excel
    
    Args:
        pages_data: 页面数据列表
        project_name: 项目名称
        
    Returns:
        Excel文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "绘本提示词"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 80
    
    # 标题行
    headers = ["页码", "画面类型", "完整提示词"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据行
    for idx, page in enumerate(pages_data, 2):
        ws.cell(row=idx, column=1, value=page.get("page_number", ""))
        ws.cell(row=idx, column=2, value=page.get("scene_type", ""))
        
        prompt_cell = ws.cell(row=idx, column=3, value=page.get("prompt", ""))
        prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # 保存文件
    export_dir = Path(settings.STORAGE_PATH) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    
    filename = f"{project_name}_prompts.xlsx"
    filepath = export_dir / filename
    
    wb.save(filepath)
    return str(filepath)


def import_prompts_from_excel(filepath: str) -> List[Dict[str, Any]]:
    """
    从Excel导入提示词
    
    Args:
        filepath: Excel文件路径
        
    Returns:
        页面数据列表
    """
    from openpyxl import load_workbook
    
    wb = load_workbook(filepath)
    ws = wb.active
    
    pages = []
    
    # 跳过标题行，从第2行开始读取
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:  # 如果页码为空，跳过
            continue
        
        pages.append({
            "page_number": str(row[0]),
            "scene_type": str(row[1]) if row[1] else "",
            "prompt": str(row[2]) if row[2] else ""
        })
    
    return pages


def export_reference_images_to_excel(
    reference_images: List[Dict[str, Any]],
    project_name: str
) -> str:
    """
    导出参考图信息到Excel
    
    Args:
        reference_images: 参考图数据列表
        project_name: 项目名称
        
    Returns:
        Excel文件路径
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "参考图提示词"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 80
    
    # 标题行
    headers = ["类型", "名称", "提示词"]
    header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 数据行
    for idx, ref_img in enumerate(reference_images, 2):
        type_map = {
            "character": "人物角色",
            "non_character": "非人物角色",
            "scene": "场景"
        }
        
        ws.cell(row=idx, column=1, value=type_map.get(ref_img.get("ref_type"), ""))
        ws.cell(row=idx, column=2, value=ref_img.get("ref_name", ""))
        
        prompt_cell = ws.cell(row=idx, column=3, value=ref_img.get("prompt", ""))
        prompt_cell.alignment = Alignment(wrap_text=True, vertical="top")
    
    # 保存文件
    export_dir = Path(settings.STORAGE_PATH) / "exports"
    export_dir.mkdir(parents=True, exist_ok=True)
    
    filename = f"{project_name}_reference_images.xlsx"
    filepath = export_dir / filename
    
    wb.save(filepath)
    return str(filepath)
